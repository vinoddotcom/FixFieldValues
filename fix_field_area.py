import openpyxl
from Neo4j.neo4j_utils import check_is_value_exist

file_name = "text_key_value_sheet_data.xlsx"
workbook = openpyxl.load_workbook(file_name)

# Define the sheet name as per your variable
field_area_subject = "Post Name--q9LzwSSI"
sheet = workbook[field_area_subject]

def update_row(sheet, row_number, new_values):
    # Ensure new_values has the same length as the number of columns in the row
    for col_num, value in enumerate(new_values, start=1):
        sheet.cell(row=row_number, column=col_num, value=value)

def read_sheet(sheet_identifier):
    if isinstance(sheet_identifier, int):
        sheet = workbook.worksheets[sheet_identifier]
    else:
        sheet = workbook[sheet_identifier]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))  # Ensure each row is a list
    
    return data


fields_data = read_sheet(field_area_subject)
for index, field in enumerate(fields_data, start=1):
    if index == 1: continue
    key_id = field[0]
    value = check_is_value_exist(key_id, field[1])
    searched_value = value[0].get("n").get("value") if value else None
    if searched_value is None:
        print(f"value: {field[2]} with key_id {field[1]} not found in Neo4j")
        continue
    elif len(value) >1:
        print(f"value: {field[2]} with key_id {field[1]} has multiple values in Neo4j")
        continue
    else:
        field[2] = searched_value
        update_row(sheet, index, field)







# # Example: Update row 3 with new values
# row_to_update = 3  # specify the row number you want to update
# new_values = ["UpdatedValue1", "UpdatedValue2", "UpdatedValue3"]  # replace with your new data

# update_row(sheet, row_to_update, new_values)

# Save changes to the workbook
workbook.save(file_name)
