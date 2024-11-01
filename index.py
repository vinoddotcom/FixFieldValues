from Neo4j.neo4j_utils import check_is_value_exist, set_value
import openpyxl


file_name = "final_text_key_value_sheet_data.xlsx"

workbook = openpyxl.load_workbook(file_name)

sheet_names = workbook.sheetnames

def read_sheet(sheet_identifier):
    if isinstance(sheet_identifier, int):
        sheet = workbook.worksheets[sheet_identifier]
    else:
        sheet = workbook[sheet_identifier]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))  # Ensure each row is a list
    return data

def fix_tag_value(key_id, value_id, corrected_value):
    existing_value = check_is_value_exist(key_id, value_id)
    if not existing_value: raise ValueError(f"key_id: {key_id}, value_id: {value_id} does't exist")
    old_value = existing_value[0].get("n").get("value")
    if corrected_value and corrected_value != old_value:
        set_value(key_id, value_id, corrected_value)
        print(key_id, value_id, corrected_value)


for sheet_name in sheet_names:
    fields_data = read_sheet(sheet_name)
    for index, field in enumerate(fields_data):
        if index == 0: continue
        key_id = field[0]
        value_id = field[1]
        corrected_value = field[3]
        fix_tag_value(key_id, value_id, corrected_value)


# fix_tag_value("eEPkh9p7","6d2pMlS3", "Post Graduate")

# from Neo4j.neo4j_utils import check_is_value_exist, set_value
# import openpyxl


# file_name = "text_key_value_sheet_data.xlsx"

# workbook = openpyxl.load_workbook(file_name)

# sheet_names = workbook.sheetnames

# def read_sheet(sheet_identifier):
#     if isinstance(sheet_identifier, int):
#         sheet = workbook.worksheets[sheet_identifier]
#     else:
#         sheet = workbook[sheet_identifier]
#     data = []
#     for row in sheet.iter_rows(values_only=True):
#         data.append(list(row))  # Ensure each row is a list
    
#     return data

# def fix_tag_value(key_id, value_id, corrected_value):
#     existing_value = check_is_value_exist(key_id, value_id)
#     if not existing_value: raise ValueError(f"key_id: {key_id}, value_id: {value_id} does't exist")
#     old_value = existing_value[0].get("n").get("value")
#     if corrected_value and corrected_value != old_value:
#         set_value(key_id, value_id, corrected_value)
#         print(key_id, value_id, corrected_value)

# def validate_sheet_data(key_id, value_id, existing_value):
#     current_value = check_is_value_exist(key_id, value_id)
#     if not current_value: 
#         print(f"key_id: {key_id}, value_id: {value_id} does't exist")
#     else:
#         val = current_value[0].get('n').get('value')
#         if not existing_value ==  val:
#             print(f"{val} === {existing_value} {key_id}")


# for sheet_name in sheet_names:
#     fields_data = read_sheet(sheet_name)
#     print(sheet_name)
#     for index, field in enumerate(fields_data):
#         if index == 0: continue
#         key_id = field[0]
#         value_id = field[1]
#         existing_value = field[2]
#         corrected_value = field[3]
#         validate_sheet_data(key_id, value_id, existing_value)


# fix_tag_value("eEPkh9p7","6d2pMlS3", "Post Graduate")