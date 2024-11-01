import json
import os
import openpyxl
from Neo4j.neo4j_utils import check_is_value_exist
from helper import update_tag_value
import tempfile
import shutil

with open("tag_value_progress.json", "r") as f:
    updated_data  = json.load(f)


file_name = "final_text_key_value_sheet_data.xlsx"

workbook = openpyxl.load_workbook(file_name)

sheet_names = workbook.sheetnames


def save_logs(new_logs, log_filename):
    existing_logs = {}
    if os.path.exists(log_filename):
        try:
            with open(log_filename, 'r') as f:
                existing_logs = json.load(f)
        except json.JSONDecodeError:
            existing_logs = {}
        except Exception as e:
            raise Exception(f"Error reading log file: {str(e)}")
    
    existing_logs.update(new_logs)
    
    tmp_dir = os.path.dirname(os.path.abspath(log_filename))
    tmp_fd, tmp_filename = tempfile.mkstemp(dir=tmp_dir, text=True)
    
    try:
        with os.fdopen(tmp_fd, 'w') as f:
            json.dump(existing_logs, f, ensure_ascii=False, indent=2)
            f.flush()
            os.fsync(f.fileno())
        
        shutil.move(tmp_filename, log_filename)
    
    except Exception as e:
        try:
            os.unlink(tmp_filename)
        except:
            pass
        raise Exception(f"Error writing log file: {str(e)}")



def read_sheet(sheet_identifier):
    if isinstance(sheet_identifier, int):
        sheet = workbook.worksheets[sheet_identifier]
    else:
        sheet = workbook[sheet_identifier]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))  # Ensure each row is a list
    return data

def fix_tag_value(key_id, value_id, corrected_value):
    if not corrected_value: return
    value = corrected_value.strip()
    if value:
        res = update_tag_value({"tagId": value_id, "tagText": value})
        updated_data[value_id] = {
                "value_id": value_id,
                "updated_response": res,
                "status": "success" if res and res== "saved" else "failed"
            }
        save_logs(updated_data, "tag_value_progress.json")
        print(key_id, value_id, value)


for sheet_name in sheet_names:
    fields_data = read_sheet(sheet_name)
    for index, field in enumerate(fields_data):
        value_id = field[1]
        if index == 0 or (value_id in updated_data and updated_data.get(value_id, {}).get("status", None) == "success"): 
            print(f"Already updated value_id: {value_id}")
            continue
        key_id = field[0]
        corrected_value = field[3]
        fix_tag_value(key_id, value_id, corrected_value)





