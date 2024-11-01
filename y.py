# import openpyxl
# from collections import defaultdict

# file_name = "text_key_value_sheet_data-v2.xlsx"

# workbook = openpyxl.load_workbook(file_name)


# def read_sheet(sheet_identifier):
#     if isinstance(sheet_identifier, int):
#         sheet = workbook.worksheets[sheet_identifier]
#     else:
#         sheet = workbook[sheet_identifier]
#     data = []
#     for row in sheet.iter_rows(values_only=True):
#         data.append(list(row))  # Ensure each row is a list
    
#     return data

# data = []
# unique_words = set()
# frequency_dict = defaultdict(list)
# word_count = defaultdict(int)
# text = ""

# def count_word_frequencies(strings):
#     for string in strings:
#         if not string: continue
#         words = string.split()
#         for word in words:
#             word_count[word] += 1
#     return word_count


# def classify_words_by_frequency():
#     for word, count in word_count.items():
#         frequency_dict[count].append(word)
#     return frequency_dict



# def add_to_text(frequency_dict, sheet_name):
#     global text
#     text += f"+++++++++++++++++++++++++++ {sheet_name} +++++++++++++++++++++++++++++\n\n"
#     for frequency, words in sorted(frequency_dict.items()):
#         text += f"{frequency} times: \n{'****'.join(words)}\n\n\n"
#     text += '\n\n\n\n\n\n'

# def save_to_file():
#     with open("unique_words-1.txt", 'w') as file:
#         file.write(text)


# sheet_names = workbook.sheetnames

# for sheet_name in sheet_names:
#     print(sheet_name)
#     data = [item[3] for item in  read_sheet(sheet_name)]
#     count_word_frequencies(data)
#     frequency_dict = classify_words_by_frequency()
#     add_to_text(frequency_dict, sheet_name)
#     data = []
#     unique_words = set()
#     frequency_dict = defaultdict(list)
#     word_count = defaultdict(int)

# save_to_file()

import openpyxl
from collections import defaultdict

file_name = "text_key_value_sheet_data.xlsx"

workbook = openpyxl.load_workbook(file_name)

def read_sheet(sheet_identifier):
    if isinstance(sheet_identifier, int):
        sheet = workbook.worksheets[sheet_identifier]
    else:
        sheet = workbook[sheet_identifier]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))  # Ensure each row is a list
    del data[0]
    return data

data = []
unique_words = set()
frequency_dict = defaultdict(list)
word_count = defaultdict(int)

def count_word_frequencies(strings):
    for string in strings:
        if not string: continue
        words = string.split()
        for word in words:
            word_count[word] += 1
    return word_count

def classify_words_by_frequency():
    for word, count in word_count.items():
        frequency_dict[count].append(word)
    return frequency_dict

def add_to_excel(worksheet, frequency_dict):
    row = 1
    worksheet.append(["Word", "Count"])  # Adding header row
    for frequency, words in sorted(frequency_dict.items(), reverse=True):
        for word in words:
            worksheet.append([word, frequency])
            row += 1

# Create a new workbook for the output
output_workbook = openpyxl.Workbook()
output_workbook.remove(output_workbook.active)  # Remove the default sheet

sheet_names = workbook.sheetnames

for sheet_name in sheet_names:
    print(sheet_name)
    data = [item[3] for item in read_sheet(sheet_name)]
    count_word_frequencies(data)
    frequency_dict = classify_words_by_frequency()
    
    # Add a new sheet to the output workbook
    output_sheet = output_workbook.create_sheet(title=sheet_name)
    add_to_excel(output_sheet, frequency_dict)
    
    data = []
    unique_words = set()
    frequency_dict = defaultdict(list)
    word_count = defaultdict(int)

# Save the output workbook
output_workbook.save("unique_words_counts.xlsx")
